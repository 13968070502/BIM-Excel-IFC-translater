"""A file to import csv files."""
import csv
import itertools
from openpyxl import load_workbook
from operator import itemgetter
import tabulate

path = "data\Input_Results_TGA_MA.csv"
file = open(path, newline= '')
reader = csv.reader(file, delimiter=';')

#header = next(reader) # The first line is the header
data = [row for row in reader] # Read the remaining data

def get_only(attribute, data):
    return [x for x in data if x['IFC_Element'] == attribute]

def show_data(data):
    """Show content of data object."""
    groups = itertools.groupby(data, key=itemgetter('IFC_Element'))
    attributes = [name for name, _ in groups]
    for attribute in attributes:
        dataset = get_only(attribute, data)
        header = dataset[0].keys()
        rows = [x.values() for x in dataset]
        print(f'{attribute}:')
        print(tabulate.tabulate(rows, header, tablefmt='grid'))



def import_from_csv(filename):
    """Read a csv file."""
    if not filename:
        filename = 'data/Input_Results_TGA_MA.csv'

    print(f'Import file from csv: {filename} ...')

    wb = load_workbook(filename)

    data = []  # list of dictionaries to catch all data

    for ws in wb.worksheets:  # loop over worksheets
        header_row = next(reader(1, 1), ws)
        print(f'Read sheet {ws.title}:')
        print(f'header: {header_row}')
        row_count = ws.max_row
        for row_idx in range(1, row_count):  # loop over rows
            cell_value = ws.cell(row=row_idx+1, column=1).value
            if not cell_value:
                continue  # skip if no data in first cell of row
            this_row_data = {}  # dictionary to catch single row content
            for col_idx, header in enumerate(header_row):  # loop over columns
                cell_value = ws.cell(row=row_idx+1, column=col_idx+1).value
                this_row_data[header] = cell_value
            data.append(this_row_data)

    show_data(data)











#for row in reader:
#
#    File_Name = str(row[0])
#    Object = str(row[1])
#    Object_Id = str(row[2])
#    IFC_Element = str(row[3])
#    Diameter = str(row[4])
#    Material = str(row[5])
#    Pipe_Color = str(row[6])
#    Piping_type = str(row[7])
#    H_V = str(row[8])
#    Total_vertices = str(row[9])
#    Xmin = str(row[10])
#    Ymin = str(row[11])
#    Xmax = str(row[12])
#    Ymax = str(row[13])
#    Project = str(row[14])
#    Building = str(row[15])
#    Floor = str(row[16])
#    Room = str(row[17])


#    data.append([File_Name,Object,Object_Id,IFC_Element,Diameter,Material,Pipe_Color,Piping_type,H_V,Total_vertices,Xmin,Ymin,Xmax,Ymax,Project,Building,Floor,Room])


#with open('data/ExcelVorlage.csv', 'r') as csv_file:
#    csv_reader = csv.DictReader(csv_file)
#
#    with open('data/ExcelVorlage_konv.csv', 'w') as new_file:
#        fieldnames = ['component', 'length', 'depth', 'height', 'material']
#
#        csv_writer = csv.DictWriter(new_file, fieldnames=fieldnames, delimiter='\t')
#
#        csv_writer.writeheader()
#
#        for line in csv_reader:
#            csv_writer.writerow(line)
"""A file to import excel files."""
import itertools
from openpyxl import load_workbook
from operator import itemgetter
import tabulate

"""A collection of utility functions for excel file handling."""


def read_in_col(coords, ws):
    """Read data from column starting below given coordinates."""
    data = []
    for row in range(coords[0], ws.max_row + 1):
        data.append(ws.cell(row=row, column=coords[1]).value)
    return data


def read_in_row(coords, ws):
    """Read data from row starting right of given coordinates."""
    data = []
    for col in range(coords[1], ws.max_column + 1):
        data.append(ws.cell(row=coords[0], column=col).value)
    return data


def get_only(component, data):
    return [x for x in data if x['COMPONENT'] == component]


def show_data(data):
    """Show content of data object."""
    groups = itertools.groupby(data, key=itemgetter('COMPONENT'))
    components = [name for name, _ in groups]
    for component in components:
        dataset = get_only(component, data)
        header = dataset[0].keys()
        rows = [x.values() for x in dataset]
        print(f'{component}:')
        print(tabulate.tabulate(rows, header, tablefmt='grid'))


def import_from_excel(filename):
    """Read an excel file."""
    if not filename:
        filename = 'data/ExcelVorlage.xlsx'

    print(f'Import file from excel: {filename} ...')

    wb = load_workbook(filename)

    data = []  # list of dictionaries to catch all data

    for ws in wb.worksheets:  # loop over worksheets
        header_row = read_in_row((1, 1), ws)
        print(f'Read sheet {ws.title}:')
        print(f'header: {header_row}')
        row_count = ws.max_row
        for row_idx in range(1, row_count):  # loop over rows
            cell_value = ws.cell(row=row_idx + 1, column=1).value
            if not cell_value:
                continue  # skip if no data in first cell of row
            this_row_data = {}  # dictionary to catch single row content
            for col_idx, header in enumerate(header_row):  # loop over columns
                cell_value = ws.cell(row=row_idx + 1, column=col_idx + 1).value
                this_row_data[header] = cell_value
            data.append(this_row_data)

    show_data(data)
